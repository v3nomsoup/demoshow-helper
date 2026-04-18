#!/usr/bin/env python3
"""Fill title / group(s) / YouTube / runtime columns in a BDSM demoshow sheet.

Reads demozoo production links from column C (as hyperlinks), fetches each page,
extracts title, group(s), and YouTube URL. Then fetches YouTube to get duration.
Writes back into columns A/B/F/G. Idempotent: only rows missing data are re-fetched.

Usage:
    python3 fill_demoshow.py [--sheet 'BDSM - April 2026'] [--file Demoshows.xlsx]

If no sheet is given, the first sheet in the workbook is used.
"""
import argparse
import json
import re
import sys
import time
import urllib.error
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import timedelta

import openpyxl

UA = ('Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 '
      '(KHTML, like Gecko) Chrome/124.0 Safari/537.36')

COL_TITLE = 1      # A
COL_GROUPS = 2     # B
COL_LINK = 3       # C (demozoo URL, used as input)
COL_YOUTUBE = 6    # F
COL_RUNTIME = 7    # G
RUNTIME_FORMAT = 'hh:mm:ss'


def fetch(url, timeout=30, attempts=5):
    last_err = None
    for i in range(attempts):
        try:
            req = urllib.request.Request(url, headers={
                'User-Agent': UA,
                'Accept-Language': 'en-US,en;q=0.9',
            })
            with urllib.request.urlopen(req, timeout=timeout) as r:
                return r.read().decode('utf-8', errors='replace')
        except urllib.error.HTTPError as e:
            last_err = e
            if e.code == 429:
                wait = 6 * (i + 1)
                print(f'    429 on {url}; sleeping {wait}s...', file=sys.stderr)
                time.sleep(wait)
                continue
            raise
        except urllib.error.URLError as e:
            last_err = e
            wait = 3 * (i + 1)
            print(f'    URLError on {url}: {e}; sleeping {wait}s...', file=sys.stderr)
            time.sleep(wait)
    raise last_err


def parse_demozoo(html):
    """Return dict(title, groups: list[str], youtube: url or None, pouet: url or None)."""
    title = None
    m = re.search(r'<div class="production_title[^"]*">\s*<h2>([^<]+)</h2>', html)
    if m:
        title = m.group(1).strip()
    if not title:
        m = re.search(r'<meta property="og:title" content="([^"]+)"', html)
        title = m.group(1).strip() if m else None

    groups = []
    m = re.search(r'<h3>\s*by\s*(.*?)</h3>', html, re.DOTALL)
    if m:
        h3 = m.group(1)
        # Groups first (most common)
        groups = re.findall(r'<a[^>]*href="/groups/\d+/"[^>]*>([^<]+)</a>', h3)
        if not groups:
            # Solo release by an individual (scener)
            groups = re.findall(r'<a[^>]*href="/sceners/\d+/"[^>]*>([^<]+)</a>', h3)
        groups = [g.strip() for g in groups]

    youtube = extract_youtube(html)

    # Pouët production link (used as fallback when demozoo has no YouTube)
    pouet = None
    m = re.search(r'https?://(?:www\.)?pouet\.net/prod\.php\?which=\d+', html)
    if m:
        pouet = m.group(0)

    return {'title': title, 'groups': groups, 'youtube': youtube, 'pouet': pouet}


def extract_youtube(html):
    """Pull the first YouTube watch URL from an HTML page."""
    # demozoo embeds the main video in a carousel JSON
    m = re.search(r'\.carousel\(\s*(\[.*?\])\s*,', html, re.DOTALL)
    if m:
        try:
            for item in json.loads(m.group(1)):
                if item.get('type') == 'video':
                    url = item.get('data', {}).get('url', '')
                    if 'youtube' in url or 'youtu.be' in url:
                        return url
        except json.JSONDecodeError:
            pass
    m = re.search(r'https?://(?:www\.)?youtube\.com/watch\?v=[A-Za-z0-9_-]+', html)
    if m:
        return m.group(0)
    m = re.search(r'https?://youtu\.be/[A-Za-z0-9_-]+', html)
    if m:
        return m.group(0)
    # Pouët sometimes only has the /embed/ form
    m = re.search(r'https?://(?:www\.)?youtube\.com/embed/([A-Za-z0-9_-]{11})', html)
    if m:
        return f'https://www.youtube.com/watch?v={m.group(1)}'
    return None


def parse_youtube_duration(html):
    """Seconds, or None."""
    m = re.search(r'"lengthSeconds":"(\d+)"', html)
    if m:
        return int(m.group(1))
    m = re.search(r'"duration":"PT(?:(\d+)H)?(?:(\d+)M)?(?:(\d+)S)?"', html)
    if m:
        h = int(m.group(1) or 0)
        mi = int(m.group(2) or 0)
        s = int(m.group(3) or 0)
        return h * 3600 + mi * 60 + s
    return None


def normalize_youtube(url):
    if not url:
        return None
    m = re.search(r'(?:v=|youtu\.be/)([A-Za-z0-9_-]{11})', url)
    if m:
        return f'https://www.youtube.com/watch?v={m.group(1)}'
    return url


def collect_rows(ws):
    """Yield (row, demozoo_url, needs_dz, needs_yt) for every row with a demozoo link."""
    for r in range(2, ws.max_row + 1):
        link_cell = ws.cell(row=r, column=COL_LINK)
        target = link_cell.hyperlink.target if link_cell.hyperlink else None
        if not target or 'demozoo.org' not in target:
            continue
        title = ws.cell(row=r, column=COL_TITLE).value
        groups = ws.cell(row=r, column=COL_GROUPS).value
        yt = ws.cell(row=r, column=COL_YOUTUBE).value
        rt = ws.cell(row=r, column=COL_RUNTIME).value
        # Re-fetch demozoo if title OR YT is missing (groups can legitimately be blank).
        needs_dz = (not title) or (not yt)
        needs_yt = bool(yt) and not rt
        yield r, target, needs_dz, needs_yt


def write_row(ws, r, dz_url, title, groups, youtube, runtime_seconds):
    if title:
        ws.cell(row=r, column=COL_TITLE, value=title)
    if groups:
        ws.cell(row=r, column=COL_GROUPS, value=' + '.join(groups))
    # Canonicalize demozoo link: cell text = URL, hyperlink = URL
    ws.cell(row=r, column=COL_LINK, value=dz_url)
    ws.cell(row=r, column=COL_LINK).hyperlink = dz_url
    yt_url = normalize_youtube(youtube)
    if yt_url:
        c = ws.cell(row=r, column=COL_YOUTUBE, value=yt_url)
        c.hyperlink = yt_url
    if runtime_seconds:
        c = ws.cell(row=r, column=COL_RUNTIME, value=timedelta(seconds=int(runtime_seconds)))
        c.number_format = RUNTIME_FORMAT


def main():
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument('--file', default='Demoshows.xlsx')
    ap.add_argument('--sheet', default=None, help='Sheet name (default: first sheet)')
    ap.add_argument('--workers', type=int, default=6,
                    help='Parallel workers for demozoo fetches (default 6)')
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.file)
    sheet_name = args.sheet or wb.sheetnames[0]
    ws = wb[sheet_name]
    print(f'Working on sheet: {sheet_name!r} in {args.file}', file=sys.stderr)

    rows = list(collect_rows(ws))
    pending = [(r, url) for r, url, needs_dz, _ in rows if needs_dz]
    print(f'{len(rows)} demozoo rows; {len(pending)} need (re)fetching', file=sys.stderr)

    # --- Phase 1: fetch demozoo pages in parallel ---
    dz_results = {}
    if pending:
        with ThreadPoolExecutor(max_workers=args.workers) as pool:
            fut = {pool.submit(fetch, url): (r, url) for r, url in pending}
            for f in as_completed(fut):
                r, url = fut[f]
                try:
                    dz_results[r] = parse_demozoo(f.result())
                    d = dz_results[r]
                    print(f'  row {r}: {d["title"]!r} by {d["groups"]} yt={d["youtube"]!r}',
                          file=sys.stderr)
                except Exception as e:
                    print(f'  row {r}: demozoo ERROR {e}', file=sys.stderr)
                    dz_results[r] = {'title': None, 'groups': [], 'youtube': None}

    # --- Phase 1.5: Pouët fallback for rows without a demozoo YouTube link ---
    pouet_to_fetch = []
    for r, dz in dz_results.items():
        if not dz.get('youtube') and dz.get('pouet'):
            pouet_to_fetch.append((r, dz['pouet']))
    if pouet_to_fetch:
        print(f'Pouët fallback: {len(pouet_to_fetch)} rows', file=sys.stderr)
        with ThreadPoolExecutor(max_workers=args.workers) as pool:
            fut = {pool.submit(fetch, url): (r, url) for r, url in pouet_to_fetch}
            for f in as_completed(fut):
                r, url = fut[f]
                try:
                    yt = extract_youtube(f.result())
                    if yt:
                        dz_results[r]['youtube'] = yt
                        print(f'  row {r}: Pouët -> yt {yt!r}', file=sys.stderr)
                    else:
                        print(f'  row {r}: Pouët had no YouTube either', file=sys.stderr)
                except Exception as e:
                    print(f'  row {r}: Pouët ERROR {e}', file=sys.stderr)

    # --- Phase 2: collect YouTube URLs to fetch ---
    # For each row, the YouTube URL may come from this fetch, or already be in the sheet.
    yt_to_fetch = []
    for r, dz_url, needs_dz, needs_yt in rows:
        if r in dz_results:
            yt_url = normalize_youtube(dz_results[r].get('youtube'))
        else:
            yt_url = normalize_youtube(ws.cell(row=r, column=COL_YOUTUBE).value)
        if yt_url:
            # Only fetch if runtime is still missing.
            if not ws.cell(row=r, column=COL_RUNTIME).value or (r in dz_results):
                yt_to_fetch.append((r, yt_url))

    yt_results = {}
    if yt_to_fetch:
        with ThreadPoolExecutor(max_workers=args.workers) as pool:
            fut = {pool.submit(fetch, url): (r, url) for r, url in yt_to_fetch}
            for f in as_completed(fut):
                r, url = fut[f]
                try:
                    s = parse_youtube_duration(f.result())
                    yt_results[r] = s
                    print(f'  row {r}: YT {url} -> {s}s', file=sys.stderr)
                except Exception as e:
                    print(f'  row {r}: YT ERROR {e}', file=sys.stderr)

    # --- Phase 3: write back ---
    for r, dz_url, needs_dz, needs_yt in rows:
        dz = dz_results.get(r, {})
        write_row(
            ws, r, dz_url,
            title=dz.get('title') or ws.cell(row=r, column=COL_TITLE).value,
            groups=dz.get('groups'),
            youtube=dz.get('youtube') or ws.cell(row=r, column=COL_YOUTUBE).value,
            runtime_seconds=yt_results.get(r),
        )

    wb.save(args.file)
    print(f'Saved {args.file}.', file=sys.stderr)

    # --- Summary ---
    total = timedelta()
    missing_yt, missing_grp = [], []
    for r in range(2, ws.max_row + 1):
        yt = ws.cell(row=r, column=COL_YOUTUBE).value
        rt = ws.cell(row=r, column=COL_RUNTIME).value
        grp = ws.cell(row=r, column=COL_GROUPS).value
        link = ws.cell(row=r, column=COL_LINK).hyperlink
        if not link:
            continue
        if isinstance(rt, timedelta):
            total += rt
        if not yt:
            missing_yt.append(r)
        if not grp:
            missing_grp.append(r)
    s = int(total.total_seconds())
    print(f'\nTotal runtime: {s // 3600}:{(s % 3600) // 60:02d}:{s % 60:02d}')
    print(f'Missing YouTube: {missing_yt}')
    print(f'Missing group(s): {missing_grp}')


if __name__ == '__main__':
    main()
