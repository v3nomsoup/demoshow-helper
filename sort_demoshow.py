#!/usr/bin/env python3
"""Sort entries in a BDSM demoshow sheet within each platform by Party Placement.

Preserves the existing platform-group order (first-appearance order in the sheet)
and sorts rows within each platform by column E (Party Placement).

Default is descending placement (worst → winner), which matches the usual
demoshow narrative of building up to the compo winner. Use --asc to flip.

Usage:
    python3 sort_demoshow.py [--sheet 'BDSM - April 2026'] [--file Demoshows.xlsx] [--asc]
"""
import argparse
import sys
from copy import copy

import openpyxl

COL_PLATFORM = 4   # D
COL_PLACEMENT = 5  # E


def snapshot_row(ws, r, ncols):
    cells = []
    for c in range(1, ncols + 1):
        cell = ws.cell(row=r, column=c)
        cells.append({
            'value': cell.value,
            'number_format': cell.number_format,
            'hyperlink': copy(cell.hyperlink) if cell.hyperlink else None,
        })
    return cells


def main():
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument('--file', default='Demoshows.xlsx')
    ap.add_argument('--sheet', default=None, help='Sheet name (default: first sheet)')
    ap.add_argument('--asc', action='store_true',
                    help='Sort placements ascending (winner first). Default is descending.')
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.file)
    sheet_name = args.sheet or wb.sheetnames[0]
    ws = wb[sheet_name]
    ncols = ws.max_column
    print(f'Sorting sheet: {sheet_name!r} in {args.file}', file=sys.stderr)

    rows = []
    for r in range(2, ws.max_row + 1):
        rows.append({
            'platform': ws.cell(row=r, column=COL_PLATFORM).value,
            'placement': ws.cell(row=r, column=COL_PLACEMENT).value,
            'cells': snapshot_row(ws, r, ncols),
        })

    platform_order = []
    for row in rows:
        if row['platform'] not in platform_order:
            platform_order.append(row['platform'])

    def place_key(p):
        if isinstance(p, (int, float)):
            return -p if not args.asc else p
        # Missing placements go last in either direction
        return float('inf')

    rows.sort(key=lambda r: (platform_order.index(r['platform']), place_key(r['placement'])))

    for i, row in enumerate(rows):
        target_row = 2 + i
        for c in range(1, ncols + 1):
            src = row['cells'][c - 1]
            cell = ws.cell(row=target_row, column=c)
            cell.value = src['value']
            cell.number_format = src['number_format']
            cell.hyperlink = src['hyperlink']

    wb.save(args.file)
    print(f'Saved {args.file}.', file=sys.stderr)


if __name__ == '__main__':
    main()
